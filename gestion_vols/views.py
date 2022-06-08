import mailbox
import pathlib
from xml.dom import NoModificationAllowedErr
from flask_mysqldb import MySQL
from flask import Flask, render_template, session, request, redirect, make_response
from .controller import functions, hashage_mdp
from .controller import bdd as bdd
import pandas, os
from werkzeug.utils import secure_filename
from openpyxl import Workbook
import json

app = Flask(__name__)
app.template_folder= "template"
app.static_folder= "static"
app.config.from_object('gestion_vols.config')
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'ienac21_bineau_dagorn_dauriac_ledergerber'
 
mysql = MySQL(app)

@app.route("/")
@app.route("/index")
@app.route("/index/<infoMsg>")
def index(infoMsg=''):
    return render_template("index.html", info=infoMsg)

@app.route("/about")
def about():
    return render_template("about.html")

@app.route("/blog")
def blog():
    return render_template("blog.html")

@app.route("/connexion")
@app.route("/connexion/<infoMsg>")
def connexion(infoMsg=''):
    return render_template("connexion.html", info=infoMsg)

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/connexion/logoutOK")

@app.route("/login", methods=["POST"])
def login():
    login = request.form['login']
    password = hashage_mdp.hash_mdp(request.form['mdp'])
    msg = functions.verifAuth(login, password)
    if "idUser" in session:
        return redirect("/index/authOK")
    else:
        return redirect("/connexion/authEchec")

@app.route("/services")
def services():
    return render_template("services.html")

@app.route("/team")
def team():
    return render_template("team.html")


@app.route("/reservation")
@app.route("/reservation/<infoMsg>")
def reservation(infoMsg=''):
    return render_template("reservation.html",info=infoMsg)

@app.route("/reservationOK")
def reservationOK():
    session.clear()
    return redirect("/reservation/reservationOK")

@app.route("/reservationPbm")
def reservationPbm():
    session.clear()
    return redirect("/reservation/reservationPbm")


@app.route("/compte")
@app.route("/compte/<infoMsg>")
def compte(infoMsg=''):
    return render_template("compte.html", info=infoMsg)

@app.route("/addUserOK")
def addUserOK():
    session.clear()
    return redirect("/compte/addUserOK")

@app.route("/addUserProblem")
def addUserProblem():
    session.clear()
    return redirect("/compte/addUserProblem")


@app.errorhandler(404)
def page_not_found(erreur):
    return make_response( render_template("404.html"), 404)

#gestion des fichiers
@app.route("/fichiers")
@app.route("/fichiers/<infoMsg>")
@app.route("/fichiers", methods=["POST"])
def fichiers(infoMsg = ''):
    if "data_excel" in request.files :
        file = request.files['data_excel']
        #enregistrement du fichier dans le repertoire files
        filename = secure_filename(file.filename)
        UPLOAD_FOLDER = os.getcwd() + "\\gestion_vols\\static\\files\\"
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        #enregistrement du fichier sur le serveur
        xls = pandas.read_excel(UPLOAD_FOLDER+file.filename)
        data = xls.to_dict('records')
        print(data)
        #enregistrement des donnees en bdd
        msg = bdd.saveDataFromFile(data)
        if msg == "addDataFromFileOK":
            return redirect("/fichiers/importDataOK")
        else :
            return redirect("/fichiers/importDataEchec")
    else :
        return render_template("/fichiers.html", info = infoMsg)


#Création de comptes
@app.route("/addMembre",methods=['POST'])
def addMembre():
    nom=request.form['NOM']
    prenom=request.form['prenom']
    mail=request.form['sonMail']
    login=request.form['login']
    motPasse=request.form['motPasse']
    statut=request.form['rad[]']
    avatar=request.form['avatar']
    msg,lastId=bdd.add_userData(nom,prenom,mail,login,motPasse,statut,avatar)
    if msg=="OK add user":
        return redirect("/compte/addUserOK")
    else:
        return redirect("/compte/addUserProblem")

#Réservation d'un créneau
@app.route("/addEvent",methods=['POST'])
def addEvent():
    
    param_idAvion_events=request.form['monSelect']
    param_text_events=request.form['comment']
    param_idTypeVol_events=request.form['rad[]']

    heure_départ=request.form['startheure']
    heure_arrivée=request.form['endheure']
    param_start_date_events=request.form['maDate'] + " " + heure_départ
    param_end_date_events=request.form['maDate'] + " " + heure_arrivée

    param_idUserReserver_events=session["idUser"]

    msg,lastId=bdd.add_eventsData(param_start_date_events, param_end_date_events, param_text_events, param_idAvion_events, param_idTypeVol_events, param_idUserReserver_events,1)
    if msg=="OK add events":
        return redirect("/reservation/reservationOK")
    else:
        return redirect("/reservation/reservationPbm")



#exporter les donnees dans un fichier
@app.route("/exportToExcel")
def exportToExcel():
    msg, listeEvents = bdd.get_tableData("events")
    wb = Workbook()
    sheet = wb.active
    
    headers = [x for x in listeEvents[[0]]]
    
    for index, value in enumerate(headers):
        sheet.cell(row=1, column= index+1).value = value
    for i, x in enumerate(listeEvents):
        for idx, value in enumerate(x.values()):
               sheet.cell(row=i+2, column=idx+1).value = value
    wb.save('myApp/static/files/export.xls')
    return redirect('/static/files/export.xls')


class DatetimeEncoder(json.JSONEncoder):
    def default(self, obj):
        try:
            return super().default(obj)
        except TypeError:
            return str(obj)


"""BACKUP
@app.route("/calendrier")
def calendrier():
    cursor = mysql.connection.cursor()
    cursor.execute('''SELECT * from `events`''')
    column_names = ('idEvent', 'start_date', 'end_date', 'text','idAvion','idTypeVol','idUserReserver','idUserEnseigner')
    evenements = cursor.fetchall()
    mysql.connection.commit()
    cursor.close()
    liste_evenements = []
    for event in evenements:
        event_to_add = {column_names[i] : event [i] for i in range(8)}
        liste_evenements.append(event_to_add)
    liste_evenements = json.dumps(liste_evenements, cls=DatetimeEncoder)
    return render_template("calendrier.html", info=liste_evenements)"""

@app.route("/calendrier")
def calendrier():

    #Récupérer les événements
    cursor = mysql.connection.cursor()
    cursor.execute('''SELECT * from `events`''')
    column_names = ('idEvent', 'start_date', 'end_date', 'text','idAvion','idTypeVol','idUserReserver','idUserEnseigner')
    evenements = cursor.fetchall()
    mysql.connection.commit()
    cursor.close()
    liste_evenements = []
    for event in evenements:
        event_to_add = {column_names[i] : event [i] for i in range(8)}
        liste_evenements.append(event_to_add)
    liste_evenements = json.dumps(liste_evenements, cls=DatetimeEncoder)

    #Récupérer la liste des avions
    liste_avions = bdd.get_tableData('`avions`')[1]
    liste_avions = json.dumps(liste_avions)

    return render_template("calendrier.html", evenements=liste_evenements, avions=liste_avions)
